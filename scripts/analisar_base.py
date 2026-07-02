#!/usr/bin/env python3
"""Analisa planilha de prospecção Instagram."""
import openpyxl

path = "D:/projetos/PROSPECCAO/Banco_Perfis_Instagram_500Plus.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

print(f"📊 PLANILHA: {path}")
print(f"Abas: {len(wb.sheetnames)}\n")

total = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
    total += count
    print(f"  {sn:35s} → {count:4d} perfis")

print(f"\n{'─' * 50}")
print(f"  TOTAL: {total} perfis em {len(wb.sheetnames)} abas")
