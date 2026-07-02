import openpyxl
wb = openpyxl.load_workbook("D:/projetos/PROSPECCAO/Banco_Perfis_Instagram_500Plus.xlsx", data_only=True)
novas_abas = ["UI UX Design", "Web Design Inspiracao", "Figma e Ferramentas", "NoCode e Webflow", "Website Designers"]
expansao = ["Landing Page Design", "Frontend Dev", "Design Studios and Agencies"]
todas = novas_abas + expansao

for sn in wb.sheetnames:
    if sn not in todas:
        continue
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        url = ws.cell(r, 5).value
        username = ws.cell(r, 2).value
        nome = ws.cell(r, 3).value
        seg = ws.cell(r, 4).value
        if url and username and str(url).startswith("https://www.instagram.com/"):
            print(f"{str(username).strip()}|{str(url).strip()}|{str(nome or '')[:40]}|{str(seg or '?')}")
