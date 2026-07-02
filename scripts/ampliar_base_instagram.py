#!/usr/bin/env python3
"""Adiciona novos perfis minerados à planilha de prospecção."""
import openpyxl, subprocess, json, sys, os, re, time
from openpyxl.styles import Font, PatternFill, Alignment

# ─── CONFIG
PLANILHA = "D:/projetos/PROSPECCAO/Banco_Perfis_Instagram_500Plus.xlsx"
VENV_ACTIVATE = "source ~/.agent-reach-venv/Scripts/activate"
OPENCLI_FLAGS = "--window background --site-session persistent"

# ─── NOVOS PERFIS (coletados nas buscas)
NOVOS = {
    "Landing Page Design": [
        ("landing.page_design", "Landing page design"),
        ("landing.page.design", ""),
        ("landing.pagedesign", "Landing page design"),
        ("hazel.landingpagedesigner", "Hazel Landing Page Designer"),
        ("webdia", "Webdia - Landing Page/design"),
        ("saaslandingpage", "SaaS Landing Page"),
        ("thisi.saaslandingpage", "Thi Si - SaaS Landing Page"),
        ("croissantsmoon", "CroissantsMoon - Landing Page SaaS"),
        ("saas.landingpage", "SaaS Landing Page Inspiration"),
        ("webuilder.gm", "Webuilder.GM - Software House"),
    ],
    "Frontend Dev": [
        ("_frontend_web_developer_", "Frontend web developer"),
        ("frontend.web.developer", "Sahil Mepani - Frontend"),
        ("frontend_web_developer_", "Frontend Web Developer"),
        ("frontend_web_developer", "Front-End Web Developer"),
        ("webticode", "WebTi Code - Front-end Developer"),
    ],
    "UI UX Design": [
        ("dashkaacom", "UI/UX Designer"),
        ("alena_mitrohina", "UI/UX Designer"),
        ("ui__ux.designer", "Mahi - UI/UX"),
        ("gautam.designn", "Gautam Arora - UI/UX"),
        ("ui.debbie", "Debbie - UI/UX Designer"),
        ("ui_portfolio", "UI Design Portfolio"),
        ("ui.fawad", "Ali Fawad Raza - UI Portfolio"),
        ("uidiksign", "UI Design Portfolio"),
        ("uidesignportfolio", "UX Designer Portfolio"),
        ("ux.ui_portfolio", "Victoria - UX/UI Design"),
        ("ux_portfolio", "UX Designer in Stockholm"),
        ("sk_visual_", "UX/UI Web Designer Portfolio"),
        ("portfolio_uxui", "UI UX Designer Portfolio"),
        ("ux_ui_portfolio", "Kate - UI/UX Designer"),
        ("jaypee_ux", "James Phuong - UX Designer"),
    ],
    "Web Design Inspiracao": [
        ("ui.mob", "Web Design Inspiration - UI Mob"),
        ("web.designinspiration", "Web Design Inspiration UI"),
        ("webdesign.ox", "Web Design Inspiration (UI/UX)"),
        ("web.inspirations", "Web Design Inspirations"),
    ],
    "Design Studios and Agencies": [
        ("web_design_agency", "Web Design Agency K&K"),
        ("web.design.agency", "Digital Marketing Agency"),
        ("webdesignagency", "Web Design Agency"),
        ("web_design_agency_", "Web Design Services"),
        ("web.design_agency", "Web Agency"),
        ("charm.digitalagency", "CHARM Digital Agency"),
        ("digital__agencyy", "Digital Agency"),
        ("mesondigital", "Digital Agency Jakarta"),
        ("freshideas.digitalagency", "Fresh Ideas Digital Agency"),
    ],
    "Figma e Ferramentas": [
        ("figma._.designer", "Creative Figma UI/UX Designer"),
        ("figma_designer_", "The Figma Designer"),
        ("figma_.designer", "Figma Designer"),
        ("figma_designer", "WEB DESIGNER"),
    ],
    "NoCode e Webflow": [
        ("webdesigner_nocode", "Web Designer No-Code"),
        ("umidjon_majiddev", "Saytlar - Web Design No-Code"),
        ("dana_webdesign", "Dana - Web Design No-code"),
        ("mypageslab", "Web Design UI UX No-code"),
        ("mgolinski_design", "Michal Golinski - UX/UI No-Code"),
    ],
    "Website Designers": [
        ("websitedesigner_ng", "Website Designer Nigeria"),
        ("websitedesigners.la", "Website Designer California"),
        ("website_designer.india", "Indian Website Designer"),
        ("website.designer_", "Website Designer"),
        ("mvmtsocials", "Website Designer - Showit & Kajabi"),
    ],
}

def get_seguidores(username):
    """Puxa followers count do Instagram via OpenCLI."""
    cmd = f'source ~/.agent-reach-venv/Scripts/activate && opencli instagram profile "{username}" -f json --window background --site-session persistent'
    r = subprocess.run(["bash", "-c", cmd], capture_output=True, text=True, timeout=15)
    if r.returncode != 0:
        return "?"
    try:
        data = json.loads(r.stdout.strip())
        if isinstance(data, list) and len(data) > 0:
            d = data[0]
            seg = d.get("followers", d.get("followers", "?"))
            return f"{seg/1000:.0f}K" if isinstance(seg, (int,float)) and seg > 1000 else str(seg)
    except:
        pass
    return "?"

def carregar_existentes(wb):
    """Carrega todos os usernames já existentes na planilha."""
    existentes = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 2).value  # coluna B = Username
            if val and isinstance(val, str) and val.strip():
                existentes.add(val.strip().lower())
    return existentes

def estilo_cabecalho(ws):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")

# ─── MAIN
print("📊 Ampliando base de prospecção...")
print(f"   Planilha: {PLANILHA}\n")

wb = openpyxl.load_workbook(PLANILHA, data_only=False)
existentes = carregar_existentes(wb)
print(f"   Perfis existentes: {len(existentes)}")

total_novos = sum(len(v) for v in NOVOS.values())
print(f"   Novos a adicionar: {total_novos}\n")

adicionados = 0
for categoria, perfis in NOVOS.items():
    # Encontrar ou criar aba
    nome_aba = categoria[:31]
    if nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        prox_linha = ws.max_row + 1
        print(f"  📁 {nome_aba} (existente, {ws.max_row - 1} perfis)")
    else:
        ws = wb.create_sheet(nome_aba)
        ws.append(["ID", "Username", "Nome", "Seguidores Est.", "Instagram URL"])
        estilo_cabecalho(ws)
        prox_linha = 2
        print(f"  📁 {nome_aba} (NOVA aba criada)")

    for username, nome in perfis:
        if username.lower() in existentes:
            continue  # já existe, pular

        # Puxar seguidores
        seguidores = get_seguidores(username)
        if seguidores == "?":
            seguidores = "—"
        
        # Adicionar
        ws.append([
            prox_linha - 1,
            username,
            nome,
            seguidores,
            f"https://www.instagram.com/{username}/"
        ])
        existentes.add(username.lower())
        adicionados += 1
        print(f"    ✅ @{username:30s} {seguidores:>8s}")
        prox_linha += 1
        time.sleep(0.5)  # pausa pra não floodar

# Salvar
try:
    wb.save(PLANILHA)
    print(f"\n{'─' * 50}")
    print(f"✅ {adicionados} novos perfis adicionados!")
    print(f"📁 Total agora: {len(existentes)} perfis em {len(wb.sheetnames)} abas")
except PermissionError:
    saida = PLANILHA.replace(".xlsx", "_NOVOS.xlsx")
    wb.save(saida)
    print(f"\n⚠️  Original estava aberto. Salvo como: {saida}")
